import openpyxl
path = r"C:\Users\aryanlak\PycharmProjects\Task\file.xlsx"
f = openpyxl.load_workbook(path)
sheet = f.active
n = []
direct = []
organic = []
paid = []
others = []
website = []
purchase = []
for cell in sheet['A']:
    n.append(str(cell.value))
for cell in sheet['B']:
    direct.append(str(cell.value))
for cell in sheet['C']:
    organic.append(str(cell.value))
for cell in sheet['D']:
    paid.append(str(cell.value))
for cell in sheet['E']:
    others.append(str(cell.value))
for cell in sheet['F']:
    website.append(str(cell.value))
for cell in sheet['G']:
    purchase.append(str(cell.value))
for j in range(len(n)):
    for k in range(j + 1, len(n)):
        if n[j] == n[k]:
            direct[j] = direct[j] + ">" + direct[k]
            organic[j] = organic[j] + ">" + organic[k]
            paid[j] = paid[j] + ">" + paid[k]
            others[j] = others[j] + ">" + others[k]
            purchase[j] = purchase[j] + purchase[k]
            n[k] = 10
wb = openpyxl.load_workbook(r"C:\Users\aryanlak\PycharmProjects\Task\Output.xlsx")
sheet = wb.active
i = 1
for j in range(1, len(n)):
    if n[j] != 10:
        p = n[j] + "-start>" + direct[j] + ">" + organic[j] + ">" + paid[j] + ">" + others[j] + ">" + website[j] + ">" + purchase[j]
        sheet.cell(row=i, column=1).value = p
        i = i+1
wb.save(r"C:\Users\aryanlak\PycharmProjects\Task\Output.xlsx")
print("Data converted")
